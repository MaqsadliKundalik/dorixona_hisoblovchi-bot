from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_excel(products, apteka, supplier, filename="hisobot.xlsx"):
    wb = Workbook()
    ws = wb.active

    # Styles
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    # Header
    ws.merge_cells("A1:L1")
    ws["A1"] = "Приложение к договору купли-продажи № ______ от __________ заключенного между"
    ws["A1"].alignment = center
    ws["A1"].font = bold

    ws["B3"] = "Поставщик:"
    ws["B3"].font = bold
    ws["C3"] = supplier
    ws["B5"] = "Получатель:"
    ws["B5"].font = bold
    ws["C5"] = apteka

    ws.merge_cells("A8:L8")
    ws["A8"] = "СПЕЦИФИКАЦИЯ"
    ws["A8"].alignment = center
    ws["A8"].font = bold

    # Table Headers
    headers = ["№", "Наименование товаров (работ, услуг)", "Ед. изм.", "Количество",
               "Отпускная цена", "Сумма", "НДС ставка", "НДС сумма",
               "Стоимость поставки с НДС", "Срок годности", "Производитель"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=col, value=header)
        cell.alignment = center
        cell.font = bold
        cell.border = border

    # Product Rows
    row_num = 13
    total_sum = total_nds = total_with_nds = 0
    for idx, prod in enumerate(products, start=1):
        suma = prod["quantity"] * prod["price"]
        nds_sum = suma * 0.12
        with_nds = suma + nds_sum
        total_sum += suma
        total_nds += nds_sum
        total_with_nds += with_nds

        row_data = [
            idx,
            prod["name"],
            "уп.",
            prod["quantity"],
            prod["price"],
            suma,
            0.12,
            nds_sum,
            with_nds,
            prod["date"].strftime("%d.%m.%Y") if isinstance(prod["date"], datetime) else prod["date"],
            prod["author"]
        ]

        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.alignment = center

        row_num += 1

    # Totals
    ws.cell(row=row_num, column=1, value="Итого на сумму").font = bold
    ws.cell(row=row_num, column=6, value=total_sum).font = bold
    ws.cell(row=row_num, column=8, value=total_nds).font = bold
    ws.cell(row=row_num, column=9, value=total_with_nds).font = bold

    row_num += 2
    ws.merge_cells(f"A{row_num}:L{row_num}")
    ws.cell(row=row_num, column=1, value=f"Всего отпущено на сумму: {total_with_nds}").font = bold

    row_num += 2
    ws.cell(row=row_num, column=1, value="Предоплата 100% от суммы").font = bold

    row_num += 2
    ws.cell(row=row_num, column=1, value=f"Всего к оплате: {total_with_nds}").font = bold

    row_num += 3
    ws.cell(row=row_num, column=2, value="Поставщик:").alignment = center
    ws.cell(row=row_num, column=2).font = bold
    ws.cell(row=row_num, column=3, value=supplier).alignment = center

    ws.cell(row=row_num, column=9, value="Получатель:").alignment = center
    ws.cell(row=row_num, column=9).font = bold
    ws.cell(row=row_num, column=10, value=apteka).alignment = center

    # Column widths adjustment
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Save workbook
    wb.save(filename)
